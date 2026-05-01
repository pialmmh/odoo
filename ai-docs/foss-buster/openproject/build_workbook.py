"""Seed the foss-buster workbook for an OpenProject slug.

Usage:
    python build_workbook.py [slug]

Defaults to slug='work-packages' if omitted. Output goes to
    ai-docs/foss-buster/openproject/<slug>/workbook.xlsx

Differences from the Odoo workbook generator:
- Adds a Phase 2.5 "Modernisation Pass" row in the Phases sheet.
- Adds a UX-Patterns overlay sheet seeded from the target-pack catalogue.
- DB-Mapping is repurposed as API-Shape-Mapping (HAL → flat DTO) and
  carries different columns to make that explicit. Sheet name kept for
  cross-pack tool compatibility.
- BFF-Plan rows are mostly proxy-and-reshape, not net-new endpoints.
- Sources lockfile keys reflect HAL API + Rails source layout.
"""
import os
import sys
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SLUG = sys.argv[1] if len(sys.argv) > 1 else 'work-packages'
OUT_DIR = f'/home/mustafa/telcobright-projects/orchestrix-v2/ai-docs/foss-buster/openproject/{SLUG}'
OUT_PATH = os.path.join(OUT_DIR, 'workbook.xlsx')
os.makedirs(OUT_DIR, exist_ok=True)

wb = Workbook()

# ── styling helpers ────────────────────────────────────────────────────
HEADER_FILL = PatternFill('solid', fgColor='1F3A5F')   # OpenProject-ish navy
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
SUBTLE_FILL = PatternFill('solid', fgColor='F0F4F8')
THIN = Side(border_style='thin', color='C0C0C0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def auto_size(ws, min_w=10, max_w=80):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        widest = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            for line in str(v).splitlines() or ['']:
                if len(line) > widest:
                    widest = len(line)
        ws.column_dimensions[letter].width = max(min_w, min(max_w, widest + 2))

def write_sheet(name, headers, rows, freeze=True):
    ws = wb.create_sheet(name) if name not in wb.sheetnames else wb[name]
    ws.append(headers)
    for c, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = BORDER
    for r in rows:
        ws.append(r)
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = BORDER
    auto_size(ws)
    if freeze:
        ws.freeze_panes = 'A2'
    return ws

wb.remove(wb.active)

# ── README ─────────────────────────────────────────────────────────────
readme = wb.create_sheet('README', 0)
readme['A1'] = 'foss-buster — clone-discovery workbook (OpenProject pack)'
readme['A1'].font = Font(bold=True, size=14)
readme.merge_cells('A1:F1')
readme['A2'] = (f'Slug: {SLUG}   |   Target: OpenProject 14 → modern React UI '
                f'(HAL-backed)   |   Generated: ' + date.today().isoformat())
readme['A2'].font = Font(italic=True, color='555555')
readme.merge_cells('A2:F2')
readme['A4'] = 'Sheet'
readme['B4'] = 'Phase'
readme['C4'] = 'Purpose'
for c in 'ABC':
    readme[f'{c}4'].fill = HEADER_FILL
    readme[f'{c}4'].font = HEADER_FONT

readme_rows = [
    ('Sources',       'P0',     'Pinned SHAs / versions / API token / dates of every input. Lockfile of the run.'),
    ('Phases',        'all',    'Gate definitions and current status per phase. Includes Phase 2.5 (Modernisation Pass).'),
    ('Stories',       'P1',     'One row per visible UI element on the source-side OpenProject screen. Pre-interaction.'),
    ('UX-Patterns',   'P2.5',   'Source pattern → modern equivalent. One row per Phase-1 story; status applied | legacy-keep | deferred.'),
    ('Interactions',  'P3',     'Each interactive element with: HAL endpoint + payload shape + lockVersion handling + form schema reference.'),
    ('DB-Mapping',    'P5',     'API-Shape-Mapping: HAL field/_links/_embedded ↔ React DTO field. Bucket: equivalent / partial / none.'),
    ('Conflict-Log',  'P4',     'Wiki↔code disagreements with resolution. Each becomes a regression test in P7.'),
    ('BFF-Plan',      'P5',     'Endpoints to add or reuse. For this target, mostly thin proxy + DTO reshape rather than net-new logic.'),
    ('Cross-Cutting', 'overlay','i18n, multi-tenancy, RBAC, workflows, attachments, notifications, saved queries.'),
    ('Glossary',      '–',      'Terms used in the columns above.'),
]
for i, row in enumerate(readme_rows, 5):
    readme[f'A{i}'] = row[0]
    readme[f'B{i}'] = row[1]
    readme[f'C{i}'] = row[2]
    for c in 'ABC':
        readme[f'{c}{i}'].alignment = Alignment(vertical='top', wrap_text=True)
        readme[f'{c}{i}'].border = BORDER
readme.column_dimensions['A'].width = 18
readme.column_dimensions['B'].width = 10
readme.column_dimensions['C'].width = 100

readme[f'A{len(readme_rows)+7}'] = 'Authority order (when sources disagree):'
readme[f'A{len(readme_rows)+7}'].font = Font(bold=True)
readme[f'A{len(readme_rows)+8}'] = ('1) HAL form payload (live)   2) OpenProject Ruby/JS source   '
                                    '3) Running OpenProject   4) Wiki + JSON   5) Screenshots / captured frames   '
                                    '6) Demo data (illustrative only)')

# ── Sources (Phase 0 lockfile) ─────────────────────────────────────────
sources_headers = ['key', 'value', 'captured_at', 'notes']
sources_rows = [
    ('slug',                       SLUG,                                                date.today().isoformat(), 'Target screen identifier'),
    ('openproject.version',        '14.6.3',                                            date.today().isoformat(), 'jammy package on Ubuntu 24.04 (per project memory)'),
    ('openproject.git_sha',        '<TODO: not a git checkout — derive from package build hash in /opt/openproject/REVISION>',
                                                                                        date.today().isoformat(), 'Reference only — source is read-only via /opt/openproject'),
    ('openproject.runtime_url',    'http://localhost:6543',                             date.today().isoformat(), 'Apache reverse proxy → Puma 127.0.0.1:6000; start via `sudo openproject-start`'),
    ('openproject.api_root',       'http://localhost:6543/api/v3',                      date.today().isoformat(), 'HAL+JSON; basic-auth (admin) or API key (Bearer/header)'),
    ('openproject.api_token',      '<TODO: generate via "My account → Access tokens"; stash here>',
                                                                                        date.today().isoformat(), 'Used by Phase-3 probes that hit the running API'),
    ('openproject.db_host',        '127.0.0.1:5433',                                    date.today().isoformat(), 'PostgreSQL 16; user=openproject, db=openproject'),
    ('react.repo',                 'orchestrix-v2/ui/',                                 date.today().isoformat(), 'Vite on :5180'),
    ('react.git_sha',              '<TODO: cd orchestrix-v2/ui && git rev-parse HEAD>', date.today().isoformat(), ''),
    ('bff.repo',                   'orchestrix-v2/api/',                                date.today().isoformat(), 'Spring Boot on :8180; APISIX :9081 — thin proxy/reshape only for this target'),
    ('bff.git_sha',                '<TODO: cd orchestrix-v2/api && git rev-parse HEAD>',date.today().isoformat(), ''),
    ('video.path',                 f'video-extractor/openproject/{SLUG}/',              date.today().isoformat(), 'Optional — if absent, capture frames live via Playwright into <slug>/captured-frames/'),
    ('wiki.path',                  f'video-extractor/wiki/openproject-{SLUG}-*.md',     date.today().isoformat(), 'Optional — same fallback as video'),
    ('demo.project_identifier',    'demo-project',                                      date.today().isoformat(), 'OpenProject ships a demo project at this slug after first install'),
    ('default_type',               '<TODO: pick first type from /api/v3/types and pin>', date.today().isoformat(), 'Stories that depend on type-specific form schemas pin this'),
]
write_sheet('Sources', sources_headers, sources_rows)

# ── Phases ─────────────────────────────────────────────────────────────
phases_headers = ['phase', 'name', 'exit_gate', 'iteration_cap', 'status', 'blockers']
phases_rows = [
    ('P0',   'Pin & Frame',
     'sources.lock written; OpenProject runtime reachable; admin API token captured; default type pinned',
     1, 'pending', ''),
    ('P1',   'Stories Table',
     'every visible field/button/affordance on the source screen has ≥1 story; confidence column populated',
     3, 'pending', ''),
    ('P2',   'Prototype UI',
     'field/intent overlap ≥ 95% (every Phase-1 story reachable in prototype); section/tab presence ≥ 90%; '
     'control-type per field matches story; visual layout MAY diverge per UX-Patterns',
     3, 'pending', ''),
    ('P2.5', 'Modernisation Pass',
     'every Phase-1 story has a UX-Patterns row; pattern_status ∈ {applied, legacy-keep, deferred}; '
     'modal-vs-page rule satisfied (no full-page edits); search input is the first affordance on every list view',
     3, 'pending', 'OpenProject pack only'),
    ('P3',   'Interaction Discovery',
     'every interactive Phase-1 row has a populated Interactions row with HAL endpoint, payload shape, '
     'and lockVersion handling noted; form-schema reference recorded for every editable field',
     3, 'pending', ''),
    ('P4',   'Reconcile',
     'zero unresolved conflicts; every resolved row has a regression test queued',
     3, 'pending', ''),
    ('P5',   'API-Shape Mapping & BFF-Plan',
     'every Interactions row has a DB-Mapping (HAL → DTO) row + a BFF-Plan row, or an explicit WONTFIX flag',
     3, 'pending', ''),
    ('P6',   'Code Lock-Step',
     'FE/BE drift ≤ 1 story; every closed story has a passing Playwright test stub generated from acceptance criteria',
     'n/a', 'pending', ''),
    ('P7',   'Replay',
     'visual diffs within tolerance (vs prototype, NOT vs OpenProject screenshots — UX has diverged); '
     'all behavioural tests green; conflict-log replayed without regression',
     'n/a', 'pending', ''),
]
write_sheet('Phases', phases_headers, phases_rows)

# ── Stories (Phase 1) ──────────────────────────────────────────────────
# Seeded with a small starter set for work-packages list/detail; agent fills
# the rest from screenshots/HAL on first P1 run.
stories_headers = [
    'story_id', 'screen', 'tab_or_section', 'control_type', 'caption',
    'actor', 'intent', 'acceptance_criteria',
    'wiki_ref', 'frame_refs', 'confidence', 'status'
]
S = lambda *r: r
stories_rows = []
if SLUG == 'work-packages':
    stories_rows = [
        S('wp.list.search',                'work-packages.list', 'header', 'search-box', 'Search work packages',
          'all', 'Filter the visible list by free-text query (matches subject + description per OpenProject `q=`)',
          'typing in the box debounces 250 ms; URL gets `?q=<term>`; results refetch; clearing the box restores prior filter set',
          '', '', 'high', 'pending'),
        S('wp.list.create_button',         'work-packages.list', 'header', 'button', '+ Work package',
          'all', 'Open the create modal pre-scoped to current project + default type',
          'click opens MUI <Dialog>; type pre-selected from project default; cancel closes without state loss; submit creates and inserts row optimistically',
          '', '', 'high', 'pending'),
        S('wp.list.filter_chips',          'work-packages.list', 'header', 'chip-row', 'Filter pills',
          'all', 'Show active filters as removable chips; click chip to edit; "+ Filter" opens picker',
          'each filter from the OpenProject saved query renders as one chip; X removes it; "+ Filter" opens a popover with field picker',
          '', '', 'high', 'pending'),
        S('wp.list.column_set',            'work-packages.list', 'list', 'columns', 'Visible columns',
          'all', 'Persist user-chosen column set; OpenProject queries.columns drives initial set',
          'first load reads /api/v3/queries/default for default columns; user changes round-trip via PATCH on the active query',
          '', '', 'high', 'pending'),
        S('wp.list.row',                   'work-packages.list', 'list', 'row', 'Work-package row',
          'all', 'Click row → open slide-over with details (modal-vs-page rule)',
          'click anywhere except a column-action opens slide-over; URL gains ?wp={id}; deep link to ?wp={id} opens slide-over on direct load',
          '', '', 'high', 'pending'),
        S('wp.list.virtualisation',        'work-packages.list', 'list', 'list-affordance', 'Virtualised infinite list',
          'all', 'Render thousands of work packages without thrashing; sticky header; resizable columns',
          'list virtualises rows beyond viewport; loads next HAL offset on scroll; row count badge reflects /api/v3/work_packages?pageSize=0 total',
          '', '', 'high', 'pending'),
        S('wp.detail.subject',             'work-packages.detail', 'header', 'inline-edit', 'Subject',
          'all', 'Click subject to inline-edit; round-trip lockVersion on save',
          'click toggles to text input; Enter saves via PATCH with lockVersion; Esc cancels; 409 surfaces "edited elsewhere" toast and refetches',
          '', '', 'high', 'pending'),
        S('wp.detail.status',              'work-packages.detail', 'header', 'status-pill', 'Status',
          'all', 'Single-select picker; available list comes from form schema (Type × Role × CurrentStatus)',
          'click pill opens menu of availableStatuses (NOT global statuses); pick status → PATCH with lockVersion; pill colour reflects new status',
          '', '', 'high', 'pending'),
        S('wp.detail.assignee',            'work-packages.detail', 'fields', 'principal-picker', 'Assignee',
          'all', 'Pick from principals (User|Group|PlaceholderUser); search-as-you-type via /api/v3/principals',
          'click opens combobox; typing queries /api/v3/principals?filters=...; picking sets _links.assignee.href; clear sets to null',
          '', '', 'high', 'pending'),
        S('wp.detail.activity',            'work-packages.detail', 'activity', 'tabbed-list', 'Activity',
          'all', 'Show journals; filter chips: All | Comments | Changes | Mentions; default Comments',
          'tab loads /api/v3/work_packages/{id}/activities; filter chips client-side; mentions chip filters to journals where current user is mentioned',
          '', '', 'high', 'pending'),
        S('wp.detail.comment_composer',    'work-packages.detail', 'activity', 'composer', 'Sticky comment composer',
          'all', 'Markdown editor at the bottom of the slide-over; @-mention via /api/v3/principals',
          'composer sticks to bottom; write/preview tabs; submit POSTs a journal with comment.raw; @-mention autocomplete inserts user-id token',
          '', '', 'high', 'pending'),
        S('wp.detail.attachments',         'work-packages.detail', 'attachments', 'drop-zone', 'Attachments drop-zone',
          'all', 'Drag files anywhere on the slide-over to attach; thumbnail grid; lightbox on click',
          'drop triggers POST /api/v3/work_packages/{id}/attachments multipart; image MIME types render thumbnail; click opens lightbox',
          '', '', 'high', 'pending'),
    ]
write_sheet('Stories', stories_headers, stories_rows)

# ── UX-Patterns (Phase 2.5 overlay) ────────────────────────────────────
# Seeded from the target-pack catalogue. Agent maps every Phase-1 story to
# one of these rows on first P2.5 run; pattern_status starts 'unset'.
ux_headers = [
    'pattern_id', 'source_pattern', 'modern_pattern', 'rationale',
    'covers_stories', 'pattern_status', 'slice_target', 'notes'
]
ux_rows = [
    ('ux.edit_modal',
     'Edit redirects to /work_packages/{id}/details (full-page).',
     'Slide-over panel from right edge; row-click opens it; URL deep-links via ?wp={id}.',
     'Linear/Notion convention; avoids context loss on lists.',
     'wp.list.row, wp.detail.*', 'unset', 'slice-1', ''),
    ('ux.create_modal',
     'Create button opens /work_packages/new full page.',
     'MUI <Dialog> with type pre-selected from current view.',
     'Faster create-flow; modal can be dismissed without losing list state.',
     'wp.list.create_button', 'unset', 'slice-1', ''),
    ('ux.cmdk_project_switcher',
     'Project switcher = sidebar dropdown.',
     'Cmd+K spotlight palette with recent + pinned + fuzzy match.',
     'Discoverable, keyboard-driven.',
     '(global — applies across slugs)', 'unset', 'slice-1', ''),
    ('ux.filter_chips',
     'Filters = horizontal toolbar above list.',
     'Chip-style filter pills under the search box; click chip to edit; "+ Filter" opens picker; saved filters in sub-menu.',
     'Compact, persistent, scannable.',
     'wp.list.filter_chips', 'unset', 'slice-1', ''),
    ('ux.virtualised_list',
     'List = paginated table with page numbers.',
     'Virtualised list with infinite scroll OR cursor pagination; sticky header; column resize.',
     'Handles thousands of rows without thrashing.',
     'wp.list.virtualisation, wp.list.column_set', 'unset', 'slice-1', ''),
    ('ux.direct_sort',
     'Sort = clicking column header opens dropdown menu.',
     'Click header to cycle asc/desc/none; modifier-click for multi-sort; sort indicators in header.',
     'Direct manipulation.',
     'wp.list.column_set', 'unset', 'slice-1', ''),
    ('ux.inline_saved_views',
     'Save query workflow uses a modal with name + visibility.',
     'Inline rename in the saved-queries menu; visibility toggle next to name; star to pin.',
     'Less ceremony.',
     '(saved queries — applies across slugs)', 'unset', 'slice-2', ''),
    ('ux.kanban_status_pills',
     'Status changes via inline single-select.',
     'Inline single-select kept, but the available list renders as kanban-pill chips coloured per status; availableStatuses from form schema.',
     'Visual state at a glance.',
     'wp.detail.status', 'unset', 'slice-1', ''),
    ('ux.sticky_composer',
     'Comment box at bottom of details page.',
     'Sticky composer at bottom of slide-over with markdown preview tab; @-mention autocomplete via /api/v3/principals.',
     'Always visible while reading the thread.',
     'wp.detail.comment_composer', 'unset', 'slice-1', ''),
    ('ux.activity_chips',
     'Activity tab = chronological list.',
     'Filter chips at top: All | Comments | Changes | Mentions; default Comments.',
     'Reduces noise.',
     'wp.detail.activity', 'unset', 'slice-1', ''),
    ('ux.dropzone_attachments',
     'Attachments = list with download links.',
     'Drop-zone over whole slide-over while dragging; thumbnail grid for images; lightbox on click.',
     'Affordance hierarchy.',
     'wp.detail.attachments', 'unset', 'slice-1', ''),
    ('ux.persistent_tree',
     'Hierarchy = nested tree with expand/collapse.',
     'Same — but expand-state persists per user in localStorage; "expand all to depth N" command.',
     'Hierarchy stays useful as data grows.',
     '(hierarchy — applies across slugs)', 'unset', 'slice-2', ''),
    ('ux.search_first',
     'Search box not always present or buried in advanced filters.',
     'Search input is the FIRST affordance on every list view; bound to OpenProject `q=`; empty = show all.',
     'Search-first nav.',
     'wp.list.search', 'unset', 'slice-1', ''),
]
write_sheet('UX-Patterns', ux_headers, ux_rows)

# ── Interactions (Phase 3) ─────────────────────────────────────────────
inter_headers = [
    'story_id', 'trigger',
    'op_summary', 'op_hal_endpoint', 'op_method',
    'op_payload_shape', 'op_form_schema_ref', 'op_lock_version_required',
    'op_journals_written', 'op_record_rules',
    'react_summary', 'react_component', 'react_query_key',
    'reshape_strategy', 'discovery_status'
]
inter_rows = []
if SLUG == 'work-packages':
    inter_rows = [
        ('wp.list.search', 'type',
         "Search debounces and re-issues GET /api/v3/work_packages with `filters=[{search:{operator:'**',values:[<q>]}}]`. Empty term clears the search filter only — other filters persist.",
         '/api/v3/work_packages',
         'GET',
         "{ filters: [{search:{operator:'**',values:[<q>]}}, ...others], pageSize, offset, sortBy }",
         '/api/v3/work_packages/form (read columns/sort defaults)',
         'no',
         '', '',
         'TanStack Query subscription keyed on (project, filters, sort, pageSize). Debounce 250 ms.',
         'WorkPackageList', '["wp", projectId, filters, sort]',
         'pass-through (HAL → DTO removes _links/_embedded; flatten _embedded.elements[]).',
         'pending'),
        ('wp.list.create_button', 'click',
         "Open MUI Dialog. POST /api/v3/work_packages/form to fetch a fresh form schema (with payload defaults). On submit, POST /api/v3/work_packages with the validated payload; lockVersion is 0 for new records.",
         '/api/v3/work_packages',
         'POST',
         "{ subject, _links:{type:{href},project:{href},status:{href},assignee:{href}?}, customField{N}? }",
         '/api/v3/work_packages/form (used to render the form fields)',
         'no (new record)',
         'one new journal (creation)', '',
         'Modal opens with form fields driven by schema; submit invalidates list query and prepends optimistic row.',
         'WorkPackageCreateDialog', '["wp", "form", projectId, typeId]',
         'reshape (schema-driven render; submit converts component state → HAL `_links` shape).',
         'pending'),
        ('wp.detail.subject', 'inline-edit',
         "Inline-edit triggers PATCH /api/v3/work_packages/{id} with { subject, lockVersion }. Server returns 200 + updated WP including new lockVersion, OR 409 with the current resource if stale.",
         '/api/v3/work_packages/{id}',
         'PATCH',
         "{ subject, lockVersion }",
         '/api/v3/work_packages/{id}/form (writable=true required)',
         'YES',
         'one journal (subject change diff)', '',
         'Inline-edit React component round-trips lockVersion. On 409: surface toast, refetch resource, drop user input (do NOT auto-merge).',
         'InlineEditField', '["wp", id]',
         'reshape (DTO → HAL has no link-fields here; pure scalar PATCH).',
         'pending'),
        ('wp.detail.status', 'select',
         "Status pill opens menu of availableStatuses (read from /api/v3/work_packages/{id}/form `_embedded.schema.status._embedded.allowedValues`). Pick → PATCH with `_links.status.href` + lockVersion.",
         '/api/v3/work_packages/{id}',
         'PATCH',
         "{ _links:{status:{href:'/api/v3/statuses/{statusId}'}}, lockVersion }",
         '/api/v3/work_packages/{id}/form (`_embedded.schema.status._embedded.allowedValues`)',
         'YES',
         'one journal (status change)', 'workflow per (Type × Role × CurrentStatus)',
         'Pill colour per status (from /api/v3/statuses); never enumerate global status list.',
         'StatusPill', '["wp", id, "form"]',
         'reshape (UI passes statusId; reshape wraps it as _links.status.href).',
         'pending'),
        ('wp.detail.assignee', 'select',
         "Combobox queries /api/v3/principals?filters=[{type_a:{operator:'=',values:['User','Group','PlaceholderUser']}}, {member:{operator:'=',values:[<projectId>]}}, {name:{operator:'~',values:[<q>]}}]. Pick → PATCH with `_links.assignee.href` + lockVersion.",
         '/api/v3/work_packages/{id}',
         'PATCH',
         "{ _links:{assignee:{href:'/api/v3/users/{userId}'|null}}, lockVersion }",
         '/api/v3/work_packages/{id}/form',
         'YES',
         'one journal (assignee change)', '',
         'PrincipalPicker uses async-select with debounce; treats Group/PlaceholderUser/User uniformly via the principal endpoint (no `users` endpoint).',
         'PrincipalPicker', '["principals", projectId, query]',
         'reshape (UI principalId+type → _links.assignee.href; clear maps to null).',
         'pending'),
        ('wp.detail.comment_composer', 'submit',
         "POST /api/v3/work_packages/{id}/activities with { comment: { raw: <markdown> } }. Server creates a journal with type=comment.",
         '/api/v3/work_packages/{id}/activities',
         'POST',
         "{ comment: { raw } }",
         '(no schema endpoint; markdown free-form)',
         'no',
         'one journal (comment)', '',
         'CommentComposer keeps draft in localStorage keyed on wp id until submit; @-mention autocomplete pulls from /api/v3/principals.',
         'CommentComposer', '["wp", id, "activities"]',
         'pass-through.',
         'pending'),
        ('wp.detail.attachments', 'drop',
         "Drop triggers POST /api/v3/work_packages/{id}/attachments multipart with the file part + metadata { description }. OpenProject also supports a 'direct upload' two-step grant for large files (request a signed URL, PUT to it, finalise).",
         '/api/v3/work_packages/{id}/attachments',
         'POST',
         "multipart/form-data: { metadata: {description}, file: <blob> }",
         '(no field schema; size/MIME limits in /api/v3/configuration)',
         'no',
         'one journal (attachment add)', '',
         'Drop-zone is the slide-over root; image MIME → thumbnail; click opens lightbox; large files use direct-upload grant.',
         'AttachmentsPanel', '["wp", id, "attachments"]',
         'pass-through (binary stays out of BFF where possible).',
         'pending'),
    ]
write_sheet('Interactions', inter_headers, inter_rows)

# ── DB-Mapping (Phase 5) — repurposed as API-Shape-Mapping ─────────────
# HAL field/_links/_embedded ↔ React DTO field. Sheet name kept for cross-pack
# tool compatibility, but columns are HAL-flavoured.
db_headers = [
    'hal_resource', 'hal_field', 'hal_kind', 'dto_field', 'dto_type',
    'mapping_bucket', 'gaps', 'notes'
]
db_rows = []
if SLUG == 'work-packages':
    db_rows = [
        ('WorkPackage', 'id',                              'scalar',          'id',                'number',                 'equivalent', '',                                                  'Stable; survives reshape'),
        ('WorkPackage', 'subject',                         'scalar',          'subject',           'string',                 'equivalent', '',                                                  ''),
        ('WorkPackage', 'lockVersion',                     'scalar',          'lockVersion',       'number',                 'equivalent', '',                                                  'Mandatory on every PATCH; UI must round-trip'),
        ('WorkPackage', '_links.status',                   'link',            'statusId',          'number',                 'partial',    'Lose href; need separate status-by-id lookup',     'reshape: extract id from href; keep statusName from _embedded.status'),
        ('WorkPackage', '_links.type',                     'link',            'typeId',            'number',                 'partial',    'Lose href',                                         'reshape: extract id'),
        ('WorkPackage', '_links.assignee',                 'link (nullable)', 'assigneeId',        'number | null',          'partial',    'Principal type lost (User|Group|PlaceholderUser)', 'Add assigneeKind discriminator on DTO'),
        ('WorkPackage', '_links.project',                  'link',            'projectId',         'number',                 'equivalent', '',                                                  ''),
        ('WorkPackage', '_embedded.status',                'embedded',        'status',            'StatusDto',              'equivalent', '',                                                  'Embedded only when ?include=status; otherwise use _links + status cache'),
        ('WorkPackage', '_embedded.assignee',              'embedded',        'assignee',          'PrincipalDto | null',    'equivalent', '',                                                  ''),
        ('WorkPackage', 'startDate',                       'scalar',          'startDate',         'iso-date | null',        'equivalent', '',                                                  ''),
        ('WorkPackage', 'dueDate',                         'scalar',          'dueDate',           'iso-date | null',        'equivalent', '',                                                  ''),
        ('WorkPackage', 'estimatedTime',                   'scalar (ISO-8601 duration)', 'estimatedHours', 'number | null', 'partial',    'Convert ISO 8601 duration → decimal hours',        'reshape: PT2H30M → 2.5'),
        ('WorkPackage', 'description.raw',                 'scalar',          'descriptionMarkdown','string',                'equivalent', '',                                                  'Description is a formattable: { raw, html, format }'),
        ('WorkPackage', 'description.html',                'scalar',          'descriptionHtml',   'string (sanitised)',     'equivalent', '',                                                  'Sanitise via DOMPurify before render'),
        ('WorkPackage', 'customField{N}',                  'scalar | link',   'cf{N}',             '<schema-derived>',       'partial',    'Type per /api/v3/work_packages/form schema',       'Schema-driven column factory; cache per (project, type)'),
        ('Collection',  '_embedded.elements',              'embedded[]',      'items',             'WorkPackageDto[]',       'equivalent', '',                                                  'List response shape'),
        ('Collection',  'total',                           'scalar',          'total',             'number',                 'equivalent', '',                                                  ''),
        ('Collection',  'pageSize',                        'scalar',          'pageSize',          'number',                 'equivalent', '',                                                  ''),
        ('Collection',  'offset',                          'scalar',          'offset',            'number',                 'equivalent', '',                                                  'Offset-based; React layer simulates cursor'),
        ('Activity',    'comment.raw',                     'scalar',          'commentMarkdown',   'string | null',          'equivalent', '',                                                  'Activity is "Journal" in the UI'),
        ('Activity',    '_links.user',                     'link',            'userId',            'number',                 'partial',    'Lose href',                                         ''),
        ('Activity',    'details[]',                       'scalar[]',        'changeDetails',     'string[]',               'equivalent', '',                                                  'Pre-rendered diff strings'),
        ('Principal',   'name',                            'scalar',          'name',              'string',                 'equivalent', '',                                                  ''),
        ('Principal',   '_type',                           'scalar',          'principalKind',     "'User'|'Group'|'PlaceholderUser'",'equivalent','',                                          'Discriminator drives icon in PrincipalPicker'),
        ('Status',      'name',                            'scalar',          'name',              'string',                 'equivalent', '',                                                  ''),
        ('Status',      'color',                           'scalar',          'color',             'string (hex)',           'equivalent', '',                                                  'Drives StatusPill colour'),
        ('Status',      'isClosed',                        'scalar',          'isClosed',          'boolean',                'equivalent', '',                                                  ''),
        ('Attachment',  'fileName',                        'scalar',          'fileName',          'string',                 'equivalent', '',                                                  ''),
        ('Attachment',  'contentType',                     'scalar',          'contentType',       'string',                 'equivalent', '',                                                  'Drives thumbnail-or-icon decision'),
        ('Attachment',  '_links.downloadLocation',         'link',            'downloadUrl',       'string',                 'equivalent', '',                                                  ''),
    ]
write_sheet('DB-Mapping', db_headers, db_rows)

# ── Conflict-Log (Phase 4) ─────────────────────────────────────────────
cf_headers = ['story_id', 'wiki_says', 'code_says', 'resolution', 'resolved_by', 'resolved_at', 'regression_test_id']
cf_rows = []
write_sheet('Conflict-Log', cf_headers, cf_rows)

# ── BFF-Plan (Phase 5) ─────────────────────────────────────────────────
# For this target, BFF rows are mostly thin proxy + reshape, NOT net-new.
bff_headers = ['endpoint', 'method', 'consumes', 'produces', 'covers_stories', 'op_strategy', 'effort', 'status']
bff_rows = []
if SLUG == 'work-packages':
    bff_rows = [
        ('/api/op/work-packages',                    'GET',    'projectId,q,filters,sort,pageSize,offset',    'WorkPackagePage{items,total,offset,pageSize}',
         'wp.list.search, wp.list.virtualisation, wp.list.column_set',
         'Proxy to /api/v3/work_packages with filter translation; reshape HAL collection → flat DTO; resolve _links by ?include= where supported',
         'small', 'pending'),
        ('/api/op/work-packages/form',               'POST',   'projectId,typeId',                            'FormSchemaDto + payloadDefaults',
         'wp.list.create_button',
         'Proxy to /api/v3/work_packages/form; flatten _embedded.schema into a render-friendly shape (one entry per editable field)',
         'small', 'pending'),
        ('/api/op/work-packages',                    'POST',   'WorkPackageCreateDto',                        'WorkPackageDto (201) | {error:validation} (400)',
         'wp.list.create_button (submit)',
         'Reshape DTO → HAL _links shape; POST /api/v3/work_packages; reshape response back to DTO',
         'small', 'pending'),
        ('/api/op/work-packages/{id}',               'GET',    'include?',                                    'WorkPackageDto',
         'wp.list.row (open slide-over via deep link), wp.detail.*',
         'Proxy GET /api/v3/work_packages/{id}; default include=status,type,assignee,project,priority',
         'small', 'pending'),
        ('/api/op/work-packages/{id}',               'PATCH',  'WorkPackagePatchDto + lockVersion',           'WorkPackageDto (200) | {error,current} (409)',
         'wp.detail.subject, wp.detail.status, wp.detail.assignee',
         'Reshape DTO → HAL; PATCH /api/v3/work_packages/{id}; propagate 409 untouched (do NOT retry / merge)',
         'small', 'pending'),
        ('/api/op/work-packages/{id}/form',          'POST',   '{ projectId?, typeId? }',                     'FormSchemaDto',
         'wp.detail.status (allowedValues), wp.detail.assignee (writable), wp.detail.subject (writable)',
         'Proxy /api/v3/work_packages/{id}/form for the modal/inline-edit affordances',
         'small', 'pending'),
        ('/api/op/work-packages/{id}/activities',    'GET',    '',                                            'JournalDto[]',
         'wp.detail.activity',
         'Proxy /api/v3/work_packages/{id}/activities; reshape Activity → JournalDto',
         'small', 'pending'),
        ('/api/op/work-packages/{id}/activities',    'POST',   '{ commentMarkdown }',                         'JournalDto (201)',
         'wp.detail.comment_composer',
         'Proxy POST /api/v3/work_packages/{id}/activities with {comment:{raw:<md>}}',
         'small', 'pending'),
        ('/api/op/work-packages/{id}/attachments',   'POST',   'multipart/form-data',                         'AttachmentDto (201)',
         'wp.detail.attachments',
         'Stream multipart through to OpenProject; or: short-circuit and return a direct-upload grant for files > 5 MB',
         'medium', 'pending'),
        ('/api/op/principals',                       'GET',    'projectId,query',                             'PrincipalDto[]',
         'wp.detail.assignee, wp.detail.comment_composer (@-mention)',
         'Proxy /api/v3/principals?filters=[type_a in {User,Group,PlaceholderUser}, member=projectId, name~query]; reshape',
         'small', 'pending'),
        ('/api/op/queries',                          'GET',    'projectId',                                   'QueryDto[]',
         '(saved views menu — applies across slugs)',
         'Proxy /api/v3/queries?filters=[project=projectId]',
         'small', 'pending'),
        ('/api/op/queries/{id}',                     'PATCH',  'QueryPatchDto',                               'QueryDto',
         '(saved views inline rename / pin)',
         'Proxy PATCH /api/v3/queries/{id}',
         'small', 'pending'),
    ]
write_sheet('BFF-Plan', bff_headers, bff_rows)

# ── Cross-Cutting overlay ──────────────────────────────────────────────
cc_headers = ['concern', 'op_mechanism', 'react_strategy', 'status']
cc_rows = [
    ('i18n',
     'Setting.default_language; per-user language; gettext-style YAML files per locale',
     'Pass through Accept-Language; read /api/v3/configuration for static labels; own dictionary for net-new affordances (filter chips, command palette)',
     'pending'),
    ('multi-tenancy',
     'None at OpenProject level — single tenant per instance; project-level isolation only',
     'Map orchestrix tenant → one OpenProject instance OR path-prefix routing (/op/{tenant}/) on the BFF; pick per deployment scale',
     'pending'),
    ('rbac',
     'Role × Project membership; per-WP form schema is authoritative for "can I edit field X right now"',
     'UI hides any control whose schema entry has writable=false; never hand-maintain a parallel permissions list',
     'pending'),
    ('workflows',
     'Status workflows: (Type × Role × CurrentStatus) → allowed_to_status; admin-editable',
     'Status selector reads availableStatuses from the form schema; never enumerate global status list',
     'pending'),
    ('auditing',
     'Journal rows per work package; exposed at /work_packages/{id}/activities',
     'Activity panel renders journals diff-style with old → new chips; read-only from React',
     'pending'),
    ('attachments',
     '/api/v3/work_packages/{id}/attachments multipart + direct-upload grants',
     'Drop-zone uploads direct to OpenProject; BFF only signs/proxies the URL grant; avoid double-buffering large files',
     'pending'),
    ('comments',
     'Journals with comment.raw (markdown)',
     'Sticky composer; @-mention autocomplete via /api/v3/principals; write/preview tabs',
     'pending'),
    ('notifications',
     '/api/v3/notifications with read flag and reason classifier',
     'Top-bar bell + slide-over notification centre; group by reason; React-Query 30s refetch',
     'pending (slice 2)'),
    ('saved-queries',
     '/api/v3/queries keyed by user; project-scoped or global',
     'Saved views menu in the list; inline rename; star to pin to sidebar',
     'pending (slice 1)'),
]
write_sheet('Cross-Cutting', cc_headers, cc_rows)

# ── Glossary ───────────────────────────────────────────────────────────
gl_headers = ['term', 'meaning']
gl_rows = [
    ('story_id',          'Stable dotted identifier; never renamed once shipped (used as test key + cross-sheet FK).'),
    ('confidence',        'How sure the agent is the row reflects reality. high = HAL+source agree; medium = source only; low = wiki implies but HAL/source disagree.'),
    ('control_type',      'Atomic UI affordance. For this pack, includes modern surfaces: search-box, chip-row, slide-over, command-palette, drop-zone, status-pill, principal-picker, sticky-composer.'),
    ('mapping_bucket',    'equivalent | partial | none — drives BFF effort estimate. Most rows here are equivalent or partial because OpenProject is the backend.'),
    ('discovery_status',  'pending | in_progress | done | wontfix — Phase 3 progress per row.'),
    ('iteration_cap',     'Max iterations the agent is allowed before stopping and surfacing to user.'),
    ('regression_test_id','File:test path that locks a Conflict-Log resolution. Empty until Phase 6.'),
    ('HAL',               'Hypertext Application Language — JSON shape OpenProject exposes at /api/v3/. Two key features: _links (URL refs to related resources) and _embedded (inlined related resources).'),
    ('lockVersion',       'OpenProject\'s optimistic-concurrency token on every editable resource. Mandatory on PATCH; PATCH without current value returns 409. UI must round-trip; BFF must propagate 409 unmodified.'),
    ('Principal',         'OpenProject\'s union of User | Group | PlaceholderUser. Assignees, watchers, responsibles all use principals — never the bare /users endpoint.'),
    ('form schema',       'Output of /api/v3/<resource>/form — authority for "can I edit field X right now" because it reflects user permissions × workflow × type.'),
    ('availableStatuses', 'The status list valid for the current (Type × Role × CurrentStatus) tuple. Read from the WP\'s form schema; NEVER from the global statuses list.'),
    ('pattern_status',    'applied | legacy-keep | deferred | unset — Phase 2.5 progress per UX-Patterns row.'),
    ('slice',             'A self-contained vertical: list+detail of one screen end-to-end. Slice 1 is "search + list + slide-over read"; slice 2 adds saved views + notifications.'),
]
write_sheet('Glossary', gl_headers, gl_rows)

# Reorder sheets so README is first.
order = ['README', 'Sources', 'Phases', 'Stories', 'UX-Patterns',
         'Interactions', 'DB-Mapping', 'Conflict-Log', 'BFF-Plan',
         'Cross-Cutting', 'Glossary']
wb._sheets = [wb[name] for name in order]

wb.save(OUT_PATH)
print(f'Wrote {OUT_PATH}')
print(f'Slug: {SLUG}')
print(f'Sheets: {", ".join(order)}')
print(f'Rows by sheet:')
for n in order:
    print(f'  {n}: {wb[n].max_row - 1}')
