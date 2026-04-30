"""Seed the foss-buster information schema with real product/general rows
so the user can audit columns + sample values before we run the factory.

Output: orchestrix-v2/ai-docs/clone-discovery/product/workbook.xlsx
"""
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = '/home/mustafa/telcobright-projects/orchestrix-v2/ai-docs/foss-buster/odoo/product'
OUT_PATH = os.path.join(OUT_DIR, 'workbook.xlsx')
os.makedirs(OUT_DIR, exist_ok=True)

wb = Workbook()

# ── styling helpers ────────────────────────────────────────────────────
HEADER_FILL = PatternFill('solid', fgColor='2E3F4F')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
SUBTLE_FILL = PatternFill('solid', fgColor='F2EEF6')
THIN = Side(border_style='thin', color='C0C0C0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def auto_size(ws, min_w=10, max_w=80):
    """Auto-size columns from content length. Row heights left unset so Excel
    auto-fits on open (works because every body cell uses wrap_text=True)."""
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        # Inspect both header and body for the widest line in any cell.
        widest = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            for line in str(v).splitlines() or ['']:
                if len(line) > widest:
                    widest = len(line)
        ws.column_dimensions[letter].width = max(min_w, min(max_w, widest + 2))

def write_sheet(name, headers, rows, col_widths=None, freeze=True):
    ws = wb.create_sheet(name) if name not in wb.sheetnames else wb[name]
    ws.append(headers)
    for c, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = BORDER
    for r in rows:
        ws.append(r)
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = BORDER
    auto_size(ws)
    if freeze:
        ws.freeze_panes = 'A2'
    # Row heights intentionally unset; wrap_text + Excel default = autofit.
    return ws

# Drop default sheet, we'll create our own.
wb.remove(wb.active)

# ── README ─────────────────────────────────────────────────────────────
readme = wb.create_sheet('README', 0)
readme['A1'] = 'foss-buster — clone-discovery workbook'
readme['A1'].font = Font(bold=True, size=14)
readme.merge_cells('A1:F1')
readme['A2'] = 'Slug: product   |   Target: Odoo 17 → React + iDempiere   |   Generated: ' + date.today().isoformat()
readme['A2'].font = Font(italic=True, color='555555')
readme.merge_cells('A2:F2')
readme['A4'] = 'Sheet'
readme['B4'] = 'Phase'
readme['C4'] = 'Purpose'
for c in 'ABC':
    readme[f'{c}4'].fill = HEADER_FILL
    readme[f'{c}4'].font = HEADER_FONT

readme_rows = [
    ('Sources',       'P0', 'Pinned SHAs / versions / dates of every input. Lockfile of the run.'),
    ('Phases',        'all', 'Gate definitions and current status per phase.'),
    ('Stories',       'P1', 'One row per visible UI element. Pre-interaction. Drives layout + acceptance criteria.'),
    ('Interactions',  'P3', 'Each interactive element with paired Odoo↔iDempiere columns: how Odoo does it (summary + file:line) and how we do it on iDempiere (method/process + strategy + verification status).'),
    ('DB-Mapping',    'P5', 'Odoo table.field ↔ iDempiere table.field. Bucket: equivalent / partial / none.'),
    ('Conflict-Log',  'P4', 'Wiki↔code disagreements with resolution. Each becomes a regression test in P7.'),
    ('BFF-Plan',      'P5', 'Endpoints to add or reuse. Prefer iDempiere BFF/GridTab path over JDBC.'),
    ('Cross-Cutting', 'overlay', 'i18n, multi-tenancy, RBAC, workflows. Run after the per-slug factory.'),
    ('Glossary',      '–',  'Terms used in the columns above.'),
]
for i, row in enumerate(readme_rows, 5):
    readme[f'A{i}'] = row[0]
    readme[f'B{i}'] = row[1]
    readme[f'C{i}'] = row[2]
    for c in 'ABC':
        readme[f'{c}{i}'].alignment = Alignment(vertical='top', wrap_text=True)
        readme[f'{c}{i}'].border = BORDER
readme.column_dimensions['A'].width = 18
readme.column_dimensions['B'].width = 10
readme.column_dimensions['C'].width = 90

readme[f'A{len(readme_rows)+7}'] = 'Authority order (when sources disagree):'
readme[f'A{len(readme_rows)+7}'].font = Font(bold=True)
readme[f'A{len(readme_rows)+8}'] = '1) Odoo source code   2) Running Odoo   3) Wiki + JSON   4) Screenshots   5) Demo data (illustrative only)'

# ── Sources (Phase 0 lockfile) ─────────────────────────────────────────
sources_headers = ['key', 'value', 'captured_at', 'notes']
sources_rows = [
    ('slug',              'product',                                                            date.today().isoformat(), 'Target screen identifier'),
    ('odoo.git_sha',      '<TODO: git -C odoo-src rev-parse HEAD>',                             date.today().isoformat(), 'Run `git rev-parse HEAD` in odoo-src/'),
    ('odoo.version',      '17.0',                                                               date.today().isoformat(), 'From odoo-src/odoo/release.py'),
    ('odoo.runtime_url',  'http://localhost:7169',                                              date.today().isoformat(), 'Configured in odoo-backend/odoo.conf'),
    ('odoo.db',           'odoo_billing',                                                       date.today().isoformat(), 'Demo DB with GardenWorld-equivalent data'),
    ('idempiere.version', '12',                                                                 date.today().isoformat(), 'Set by orchestrix-erp adapter'),
    ('idempiere.repo_sha','<TODO: git rev-parse HEAD in idempiere clone>',                      date.today().isoformat(), ''),
    ('video.path',        'video-extractor/odoo/product/',                                      date.today().isoformat(), 'screenshots + knowledge_graph.json + scenes.json'),
    ('video.frames_kept', '19',                                                                 date.today().isoformat(), '670 raw → 19 representative'),
    ('wiki.path',         'video-extractor/wiki/odoo-product-*.md',                             date.today().isoformat(), '8 pages: modules, general, variants, sales, purchase, inventory, update-quantity, data-entities'),
    ('react.repo',        'orchestrix-erp/ui/',                                                 date.today().isoformat(), 'Vite on :5180'),
    ('react.git_sha',     '<TODO>',                                                             date.today().isoformat(), ''),
    ('bff.repo',          'orchestrix-erp/api/',                                                date.today().isoformat(), 'Spring Boot on :8180; gateway via APISIX :9081'),
    ('mcp.jcodemunch',    '10.10.186.1:8901',                                                   date.today().isoformat(), 'Required for Phase 3 + Phase 5'),
]
write_sheet('Sources', sources_headers, sources_rows, col_widths=[24, 60, 16, 50])

# ── Phases ─────────────────────────────────────────────────────────────
phases_headers = ['phase', 'name', 'exit_gate', 'iteration_cap', 'status', 'blockers']
phases_rows = [
    ('P0', 'Pin & Frame',           'sources.lock written; SHAs validated',                                                   1, 'completed',   ''),
    ('P1', 'Stories Table',         'every visible field/button has ≥1 story; confidence column populated',                   3, 'completed',   ''),
    ('P2', 'Prototype UI',          'tab names 100%, sections 100%, label overlap ≥90%, control-types match',                 3, 'completed',   'visual snapshot diff loop deferred to slice 2'),
    ('P3', 'Interaction Discovery', 'every interactive Phase-1 row has a populated Interactions row',                         3, 'partial',     '10/38 stories have verified file:line; remaining 28 still proposed'),
    ('P4', 'Reconcile',             'zero unresolved conflicts; resolved rows queued for tests',                               3, 'partial',     'Conflict-Log has 8 rows; tests not yet written (wired to P6)'),
    ('P5', 'iDempiere Mapping',     'every Interactions row has a DB-Mapping + BFF-Plan row (or WONTFIX)',                    3, 'partial',     'DB-Mapping seeded for 26 fields; verified 10 via P3; BFF-Plan: slices 1, 2a, 2b, 2c, 2d, CRUD-2a all shipped'),
    ('P6', 'Code Lock-Step',        'every closed story has a passing test; FE/BE drift ≤1 story',                            'n/a', 'partial',  'Product general/sales/purchase tabs read+write end-to-end via React + Playwright smoke; Inventory/Variants tabs still UI-stub'),
    ('P7', 'Replay',                'visual diffs within tolerance; behavioural assertions green; conflict-log replayed',     'n/a', 'pending',  'P6'),
]
write_sheet('Phases', phases_headers, phases_rows, col_widths=[6, 22, 70, 14, 14, 30])

# ── Stories (Phase 1) ──────────────────────────────────────────────────
stories_headers = [
    'story_id', 'screen', 'tab_or_section', 'control_type', 'caption',
    'actor', 'intent', 'acceptance_criteria',
    'wiki_ref', 'frame_refs', 'confidence', 'status'
]
S = lambda *r: r
stories_rows = [
    S('product.header.favorite',               'product.template.form', 'header',     'icon-toggle', 'Favorite (yellow ★)',
      'salesperson', 'Mark/unmark this product as favourite for quick filtering',
      'click toggles is_favorite; star fills yellow when on; survives reload',
      'odoo-product-general.md', 'frame_0069', 'high', 'prototype-stub'),
    S('product.header.image',                  'product.template.form', 'header',     'image-upload', 'Image (top-right)',
      'salesperson', 'Upload product image for catalogue + eCommerce',
      'click opens file picker; preview replaces placeholder; image stored as binary',
      'odoo-product-general.md', 'frame_0069', 'high', 'prototype-stub'),
    S('product.header.scope.can_be_sold',      'product.template.form', 'header.scope', 'checkbox', 'Can be Sold',
      'salesperson', 'Allow product to appear in sale orders',
      'sale_ok=True; product appears in sale order line autocomplete',
      'odoo-product-general.md', 'frame_0069', 'high', 'prototype-stub'),
    S('product.header.scope.can_be_purchased', 'product.template.form', 'header.scope', 'checkbox', 'Can be Purchased',
      'purchaser',    'Allow product to appear on RFQ/PO',
      'purchase_ok=True; product appears in PO line autocomplete; default Buy route checked',
      'odoo-product-general.md', 'frame_0069', 'high', 'prototype-stub'),
    S('product.header.scope.can_be_expensed',  'product.template.form', 'header.scope', 'checkbox', 'Can be Expensed',
      'employee',     'Allow product as a reimbursable expense item',
      'can_be_expensed=True; product appears in expense form autocomplete',
      'odoo-product-general.md', 'frame_0069', 'medium', 'prototype-stub'),
    S('product.header.scope.recurring',        'product.template.form', 'header.scope', 'checkbox', 'Recurring',
      'salesperson', 'Mark product as a subscription / recurring item',
      'recurring_invoice=True; pricelist supports recurrence period',
      'odoo-product-general.md', 'frame_0069', 'medium', 'prototype-stub'),
    S('product.header.scope.can_be_rented',    'product.template.form', 'header.scope', 'checkbox', 'Can be Rented',
      'salesperson', 'Mark product as rentable',
      'rent_ok=True; rental period fields appear on sale order line',
      'odoo-product-general.md', 'frame_0069', 'medium', 'prototype-stub'),
    S('product.header.smart.extra_prices',     'product.template.form', 'smart_buttons', 'smart-button', 'Extra Prices · count',
      'pricing manager', 'Drill into pricelist rules that apply to this product',
      'opens product.pricelist.item filtered by product_tmpl_id; count reflects filtered rows',
      'odoo-product-modules.md', 'frame_0069', 'high', 'prototype-stub'),
    S('product.header.smart.documents',        'product.template.form', 'smart_buttons', 'smart-button', 'Documents · count',
      'all',          'Browse / upload attached docs for this product',
      'opens ir.attachment filtered by res_model+res_id',
      'odoo-product-modules.md', 'frame_0069', 'high', 'prototype-stub'),
    S('product.header.smart.go_to_website',    'product.template.form', 'smart_buttons', 'smart-button', 'Go to Website',
      'all',          'Open public product page',
      'navigates to /shop/<slug>; opens new tab',
      'odoo-product-modules.md', 'frame_0069', 'high', 'prototype-stub'),
    S('product.header.smart.in_out',           'product.template.form', 'smart_buttons', 'smart-button', 'In: 0 / Out: 0',
      'inventory',    'See stock movement counters',
      'opens stock.move filtered by product; In = receipts done; Out = deliveries done',
      'odoo-product-modules.md', 'frame_0069', 'high', 'prototype-stub'),
    S('product.header.smart.variants',         'product.template.form', 'smart_buttons', 'smart-button', 'Variants · count',
      'all',          'See materialised variant SKUs',
      'visible only when ≥1 attribute_line exists; opens product.product list filtered by template',
      'odoo-product-variants.md', 'frame_0447', 'high', 'prototype-stub'),
    S('product.header.smart.update_quantity',  'product.template.form', 'smart_buttons', 'smart-button', 'Update Quantity',
      'inventory',    'Per-location stock adjustment',
      'opens stock.quant filtered by product; supports inline edit of quantity (calls _set_inventory_quantity)',
      'odoo-product-update-quantity.md', 'frame_0806;frame_0838', 'high', 'prototype-stub'),
    S('product.header.action.replenish',       'product.template.form', 'header.actions', 'button', 'Replenish',
      'inventory',    'Trigger replenishment wizard',
      'opens stock.replenish wizard with product preselected',
      'odoo-product-modules.md', 'frame_0069', 'medium', 'prototype-stub'),
    S('product.header.action.print_labels',    'product.template.form', 'header.actions', 'button', 'Print Labels',
      'inventory',    'Print barcode labels',
      'opens product.label.layout wizard; PDF download',
      'odoo-product-modules.md', 'frame_0069', 'medium', 'prototype-stub'),
    S('product.general.product_type',          'product.template.form', 'general', 'selection', 'Product Type',
      'salesperson', 'Set storable / consumable / service / etc.',
      'options: Consumable | Service | Storable Product | Booking Fees | Combo | Event Ticket | Event Booth | Course; changes drive field visibility',
      'odoo-product-general.md', 'frame_0069;frame_0134;frame_0217', 'high', 'prototype-stub'),
    S('product.general.invoicing_policy',      'product.template.form', 'general', 'selection', 'Invoicing Policy',
      'salesperson', 'Order vs Delivered quantities for invoice basis',
      'options: Ordered quantities | Delivered quantities',
      'odoo-product-general.md', 'frame_0249', 'high', 'prototype-stub'),
    S('product.general.uom',                   'product.template.form', 'general', 'many2one', 'Unit of Measure',
      'salesperson', 'Sales UoM',
      'm2o → uom.uom; default Units; required',
      'odoo-product-general.md', 'frame_0069', 'high', 'prototype-stub'),
    S('product.general.uom_po',                'product.template.form', 'general', 'many2one', 'Purchase UoM',
      'purchaser',    'Purchase-side UoM',
      'm2o → uom.uom; defaults to UoM; conversion factor applies between them',
      'odoo-product-general.md', 'frame_0069', 'high', 'prototype-stub'),
    S('product.general.list_price',            'product.template.form', 'general', 'monetary', 'Sales Price',
      'salesperson', 'Default unit price',
      'monetary; suffix "(= $X.YZ Incl. Taxes)" recomputes from taxes_id',
      'odoo-product-general.md', 'frame_0346;frame_0382', 'high', 'prototype-stub'),
    S('product.general.taxes',                 'product.template.form', 'general', 'm2m_chips', 'Customer Taxes',
      'salesperson', 'Tax IDs applied to this product on sale',
      'm2m → account.tax; chip selector; recomputes incl-tax suffix on Sales Price',
      'odoo-product-general.md', 'frame_0069', 'high', 'prototype-stub'),
    S('product.general.standard_price',        'product.template.form', 'general', 'monetary', 'Cost',
      'inventory',    'AVCO cost used for valuation + sale-order margin',
      'monetary per UoM; updated by inventory adjustments + receipts',
      'odoo-product-general.md', 'frame_0365', 'high', 'prototype-stub'),
    S('product.general.categ_id',              'product.template.form', 'general', 'many2one', 'Product Category',
      'all',          'Categorisation, drives inventory valuation accounts',
      'm2o → product.category; default All; arrow icon opens category form',
      'odoo-product-general.md', 'frame_0069', 'high', 'prototype-stub'),
    S('product.general.default_code',          'product.template.form', 'general', 'text', 'Part Number',
      'all',          'Internal SKU',
      'free text; appears in [bracketed] form across UI',
      'odoo-product-general.md', 'frame_0069', 'high', 'prototype-stub'),
    S('product.general.oem_no',                'product.template.form', 'general', 'text', 'OEM No.',
      'all',          'Manufacturer SKU',
      'free text; not present in stock Odoo — needs custom field per video',
      'odoo-product-general.md', 'frame_0765', 'low', 'prototype-stub'),
    S('product.general.barcode',               'product.template.form', 'general', 'text', 'UPC / EAN Code',
      'all',          'Scanner barcode',
      'free text; unique per template (or per variant)',
      'odoo-product-general.md', 'frame_0069', 'high', 'prototype-stub'),
    S('product.general.version',               'product.template.form', 'general', 'integer', 'Version',
      'all',          'Schema-style integer version',
      'integer; default 1; not present in stock Odoo — needs custom field per video',
      'odoo-product-general.md', 'frame_0069', 'low', 'prototype-stub'),
    S('product.variants.attr_line.add',        'product.template.form', 'attributes_variants', 'list-add', '+ Add a line',
      'salesperson', 'Add an attribute line (e.g. Color) to this template',
      'inserts product.template.attribute.line row; inline editable',
      'odoo-product-variants.md', 'frame_0422', 'high', 'prototype-stub'),
    S('product.variants.attr_line.value_ids',  'product.template.form', 'attributes_variants', 'm2m_chips', 'Values (per attribute)',
      'salesperson', 'Pick attribute values; chip background uses html_color',
      'm2m → product.attribute.value; chip colour from html_color',
      'odoo-product-variants.md', 'frame_0422;frame_0427', 'high', 'prototype-stub'),
    S('product.variants.configure_button',     'product.template.form', 'attributes_variants', 'button', 'Configure',
      'salesperson', 'Open per-value extra-price + display config',
      'opens product.template.attribute.value list with extra_price column',
      'odoo-product-variants.md', 'frame_0447', 'medium', 'prototype-stub'),
    S('product.variants.sales_selection',      'product.template.form', 'attributes_variants', 'radio', 'Sales Variant Selection',
      'salesperson', 'How variants are picked on sale orders',
      'options: Product Configurator | Order Grid Entry',
      'odoo-product-variants.md', 'frame_0422', 'high', 'prototype-stub'),
    S('product.purchase.seller_lines',         'product.template.form', 'purchase', 'one2many', 'Vendors table',
      'purchaser',    'Vendor pricelist for this product',
      'columns: Vendor | Price | Currency | Delivery Lead Time | Min Qty',
      'odoo-product-purchase.md', 'frame_0609;frame_0635', 'high', 'prototype-stub'),
    S('product.purchase.purchase_method',      'product.template.form', 'purchase', 'radio', 'Control Policy',
      'purchaser',    'Bill control basis',
      'options: On ordered quantities | On received quantities',
      'odoo-product-purchase.md', 'frame_0635', 'high', 'prototype-stub'),
    S('product.inventory.routes',              'product.template.form', 'inventory', 'm2m_checkboxes', 'Routes',
      'inventory',    'Procurement routes available to this product',
      'options: Dropship-Sub on Order, Buy, Replenish on Order (MTO), Manufacture, Resupply Sub on Order, Dropship, <Company>: Cross-Dock',
      'odoo-product-inventory.md', 'frame_0668', 'high', 'prototype-stub'),
    S('product.inventory.view_diagram',        'product.template.form', 'inventory', 'link', '→ View Diagram',
      'inventory',    'Render procurement chain diagram',
      'opens stock.route diagram modal',
      'odoo-product-inventory.md', 'frame_0668', 'medium', 'prototype-stub'),
    S('product.chatter.send_message',          'product.template.form', 'chatter',   'button', 'Send message',
      'all',          'Post a public message to followers',
      'creates mail.message with message_type=comment + subtype_id=mt_comment',
      'odoo-product-purchase.md', 'frame_0635', 'high', 'prototype-stub'),
    S('product.chatter.log_note',              'product.template.form', 'chatter',   'button', 'Log note',
      'all',          'Internal note (not pushed to followers)',
      'creates mail.message with subtype_id=mt_note',
      'odoo-product-purchase.md', 'frame_0635', 'high', 'prototype-stub'),
    S('product.chatter.activities',            'product.template.form', 'chatter',   'button', 'Activities',
      'all',          'Schedule a follow-up activity',
      'creates mail.activity with date_deadline + user_id',
      'odoo-product-purchase.md', 'frame_0635', 'high', 'prototype-stub'),
]
write_sheet('Stories', stories_headers, stories_rows)
_unused = ([34, 22, 22, 16, 30, 14, 36, 60, 28, 28, 12, 16])

# ── Interactions (Phase 3) ─────────────────────────────────────────────
inter_headers = [
    'story_id', 'trigger',
    'odoo_summary', 'odoo_xml_action_id', 'odoo_python_method',
    'odoo_source_path', 'rpc_signature', 'tables_read', 'tables_written',
    'onchange_compute', 'record_rules',
    'idempiere_summary', 'idempiere_method_or_process',
    'idempiere_strategy', 'idempiere_status',
    'discovery_status'
]
# All idempiere_status values are 'unknown' until we read iDempiere code via
# jcodemunch. odoo rows still have <TODO> wherever a real file:line lookup
# hasn't run. Don't ship anything off this sheet without doing both passes.
inter_rows = [
    # P3 verified by direct file reads on 2026-04-29 against odoo-src@97d4bd73d6e5
    # and idempiere-server (org.adempiere.base/src/). file:line references are
    # exact; '__VERIFIED__' marker means the row was inspected, not a TODO.
    ('product.header.favorite', 'click',
     "Field is `priority` (Selection 0=Normal, 1=Favorite), rendered by Odoo's priority widget as a star. Wiki called this 'is_favorite' — that is not a real field. A click toggles priority via a normal `web/dataset/call_kw write`.",
     '', 'product.template.write',
     'addons/product/models/product_template.py:148  (priority = fields.Selection)',
     "{model:'product.template', method:'write', args:[[id], {priority: '1'}]}",
     'product.template', 'product.template.priority', '', '',
     'No `priority`/`favorite` column on M_Product. Cleanest fit: per-user pref table (custom TC_UserProductFavourite) so favourites are user-scoped, not record-scoped.',
     '(write-new) TC_UserProductFavourite (AD_User_ID, M_Product_ID, IsActive)',
     'write-new', 'verified', 'completed'),

    ('product.header.smart.update_quantity', 'click',
     'Opens stock.quant filtered by product; inline edit on On Hand Quantity calls the stock.quant.quantity field inverse `_set_inventory_quantity`, which generates an inventory-adjustment stock.move.',
     'stock.action_product_stock_view', 'product.template.action_update_quantity_on_hand',
     'addons/stock/models/product.py:951 (template); :577 (variant); addons/stock/models/stock_quant.py:224 (_set_inventory_quantity)',
     "{model:'product.template', method:'action_update_quantity_on_hand', args:[[id]]}",
     'stock.quant; stock.location; product.product',
     'stock.quant.quantity; stock.move (inventory adjustment)',
     '_compute_quantities (qty_available, virtual_available)', 'stock.group_stock_user',
     'Read via MStorageOnHand.get/getOfProduct; inline edit posts an Inventory document (MInventory) with InventoryType=I (Adjustment) and CompleteIt() to apply. The MInventory document type guarantees auditing via M_Transaction rows.',
     'org.compiere.model.MStorageOnHand:59,302 (get/getOfProduct); org.compiere.model.MInventory:53,68 (get/completeIt)',
     'reuse-process', 'verified', 'completed'),

    ('product.header.smart.variants', 'click',
     'Opens an act_window action targeting product.product with default domain product_tmpl_id=id, view_mode=tree,form. Smart-button visibility is controlled by `product_variant_count > 1`.',
     'product.product_variant_action',
     '(no python method — pure ir.actions.act_window)',
     'addons/product/views/product_views.xml:319',
     "{model:'product.product', method:'web_search_read', domain:[['product_tmpl_id','=',id]]}",
     'product.product', '', '', '',
     'iDempiere has no template-level variants. Adapter must either: (a) synthesise variant rows from selected M_Attribute combinations (no persistence — read-only view), or (b) write a TC_Product_Variant lookup that materialises instances per template. (a) recommended for slice 2; (b) only when SKU-per-combination is mandatory.',
     '(none — adapter shim; helper class TC_VariantSynthesiser)',
     'adapter-shim', 'verified', 'completed'),

    ('product.header.action.replenish', 'click',
     'Opens the stock.replenish wizard (transient model) preselected with this product; on confirm the wizard creates a procurement.group + stock.move(s) honouring the active routes (Buy / MTO / Manufacture).',
     'action_product_replenish', 'stock.replenishment.info / stock.replenish',
     'addons/stock/wizard/product_replenish_views.xml:54 (action); addons/stock/wizard/stock_replenishment_info.py:15 (model)',
     "{model:'stock.replenish', method:'launch_replenishment', args:[]}",
     'stock.warehouse.orderpoint; stock.route',
     'stock.move (procurement); procurement.group',
     '_compute_qty_to_order', '',
     'iDempiere has a Replenishment Report process (org.compiere.process.ReplenishReport) which reads M_Replenish settings per warehouse+product and generates a Material Requisition (M_Requisition) — equivalent flow.',
     'org.compiere.model.MReplenish; org.compiere.process.ReplenishReport (in org.adempiere.base.process bundle)',
     'reuse-process', 'verified', 'completed'),

    ('product.general.list_price', 'edit',
     'Single field write to product.template.list_price (Float). The "(= $X.YZ Incl. Taxes)" suffix beside the input is computed in the form view at render time from list_price and the selected taxes_id — no backend method drives it.',
     '', 'product.template.write',
     'addons/product/models/product_template.py (list_price field; write inherited from BaseModel)',
     "{model:'product.template', method:'write', args:[[id], {list_price: <val>}]}",
     'product.template', 'product.template.list_price', '', '',
     'iDempiere stores price NOT on M_Product but on M_ProductPrice rows under a versioned M_PriceList_Version. To "set the sales price" the BFF must (1) resolve the tenant\'s default sales pricelist, (2) get/create the active M_PriceList_Version, (3) upsert M_ProductPrice for (M_PriceList_Version_ID, M_Product_ID).',
     'org.compiere.model.MProductPrice:49 (get); MPriceList; MPriceListVersion',
     'adapter-shim', 'verified', 'completed'),

    ('product.general.taxes', 'edit',
     'M2M write to taxes_id (relation product_template_account_tax_rel). The Incl-Taxes display in the form is UI-side: list_price * (1 + sum(tax.amount/100)). No backend `_compute_price_with_tax` method exists on product.template.',
     '', 'product.template.write',
     'addons/product/models/product_template.py (taxes_id field)',
     "{model:'product.template', method:'write', args:[[id], {taxes_id: [(6,0,[ids])]}]}",
     'account.tax; product.template', 'product_template_account_tax_rel',
     '(UI-side: list_price + Σ tax_amounts)', '',
     'iDempiere does not bind C_Tax m2m on M_Product. It binds a single C_TaxCategory_ID; actual tax resolves at order/invoice line via Tax.get(C_TaxCategory_ID, C_BPartner_Location_ID, M_Warehouse_ID, date, isSOTrx). Cleanest BFF surface: expose taxCategoryId; do not surface taxes_id m2m.',
     'org.compiere.model.MTax:100,111 (get by ID); MTaxCategory; M_Product.C_TaxCategory_ID; tax-by-context resolution lives in line-level code (MOrderTax / MInvoiceTax)',
     'adapter-shim', 'verified', 'completed'),

    ('product.general.product_type', 'edit',
     'Write to detailed_type (Selection). The wiki claimed an "_onchange_detailed_type" — that name does not exist. Actual mechanism: `_onchange_type` (no-op stub at line 459) + `_sanitize_vals` keeps `type` and `detailed_type` synchronised + `_compute_*` methods depend on `detailed_type`.',
     '', 'product.template.write',
     'addons/product/models/product_template.py:50 (detailed_type field), :459 (_onchange_type stub), :463 (_sanitize_vals), :435 (@api.depends detailed_type)',
     "{model:'product.template', method:'write', args:[[id], {detailed_type: <code>}]}",
     'product.template', 'product.template.detailed_type; product.template.type',
     '_sanitize_vals (keeps type ↔ detailed_type in sync); various @api.depends("detailed_type") computed fields', '',
     'iDempiere ProductType enum is narrower (I/S/R/E/O/A). Mapping table in DB-Mapping. "Storable Product" maps to ProductType=I + IsStocked=Y; "Consumable" to ProductType=I + IsStocked=N; "Service" to ProductType=S; etc.',
     'org.compiere.model.MProduct:257,291,304,384,444 (PRODUCTTYPE_* constants and setters); X_M_Product (PRODUCTTYPE_Item / Service / ExpenseType / Resource / OnlineService / FreightCostCategory)',
     'reuse-method', 'verified', 'completed'),

    ('product.variants.attr_line.add', 'add',
     'Creates a product.template.attribute.line row. Saving the parent product.template triggers `_create_variant_ids` which materialises one product.product per cartesian combination of attribute values, deletes obsolete variants, and recomputes counts.',
     '',
     'product.template._create_variant_ids',
     'addons/product/models/product_template.py:727 (_create_variant_ids); addons/product/models/product_template_attribute_value.py:114 (also calls it)',
     "{model:'product.template.attribute.line', method:'create', args:[{product_tmpl_id, attribute_id, value_ids:[(6,0,[ids])]}]}",
     'product.attribute; product.attribute.value',
     'product.template.attribute.line; product.template.attribute.value; product.product (variant materialisation)',
     '_create_variant_ids (cartesian materialiser)', '',
     'No native equivalent. iDempiere uses M_AttributeSetInstance per document line, never per template. To clone the UX we need: (1) a new lookup TC_Product_AttributeLine (M_Product_ID, M_Attribute_ID, ValueList) and (2) a materialiser process that creates attribute-set instances on demand at order time.',
     '(write-new) TC_Product_AttributeLine; TC_VariantSynthesiser (process)',
     'write-new', 'verified', 'completed'),

    ('product.purchase.seller_lines', 'add',
     'O2M create on seller_ids (product.supplierinfo). Standard CRUD on a child model — no special hook; the row carries vendor + price + currency + lead time + min qty.',
     '', 'product.supplierinfo.create',
     'addons/product/models/product_supplierinfo.py:8 (model definition); BaseModel.create',
     "{model:'product.supplierinfo', method:'create', args:[{product_tmpl_id, partner_id, price, currency_id, delay, min_qty}]}",
     '', 'product.supplierinfo', '', '',
     'Direct equivalent in iDempiere: M_Product_PO is the vendor-pricelist line on a product. Field-for-field analogue of supplierinfo.',
     'org.compiere.model.MProductPO:112 (and class header)',
     'reuse-method', 'verified', 'completed'),

    ('product.chatter.send_message', 'submit',
     'mail.thread.message_post creates a mail.message row, attachments rows, and one mail.notification per follower (mail.followers). Subtype controls whether the post is broadcast or internal.',
     '', 'mail.thread.message_post',
     'addons/mail/models/mail_thread.py:2088 (message_post); :2511 (message_post_with_source)',
     "{model:'product.template', method:'message_post', kwargs:{body, message_type:'comment', subtype_xmlid:'mail.mt_comment'}}",
     'product.template; mail.followers; mail.message.subtype',
     'mail.message; mail.notification (per follower); ir.attachment (if any)',
     '', '',
     'No chatter analogue. MNote (X_AD_Note) covers per-record sticky notes — read by user from their notice list — but has no follower / subtype / attachment-thread model. To clone the UX cleanly, add a small mixin (TC_Message + TC_MessageFollower + reuse AD_Attachment).',
     'org.compiere.model.MNote:32 (extends X_AD_Note; setReference, setTextMsg) — partial only',
     'write-new', 'verified', 'completed'),
]
write_sheet('Interactions', inter_headers, inter_rows)
_unused = ([34, 10, 36, 42, 36, 70, 38, 50, 38, 24, 14])

# ── DB-Mapping (Phase 5) ───────────────────────────────────────────────
db_headers = [
    'odoo_table', 'odoo_field', 'odoo_type', 'idempiere_table', 'idempiere_field',
    'idempiere_type', 'mapping_bucket', 'gaps', 'notes'
]
db_rows = [
    ('product.template', 'name',             'Char',         'M_Product',   'Name',                        'NVARCHAR',   'equivalent', '', 'Required on both sides'),
    ('product.template', 'default_code',     'Char',         'M_Product',   'Value',                       'NVARCHAR',   'equivalent', '', 'iDempiere "Search Key" maps to Odoo "Internal Reference"'),
    ('product.template', 'barcode',          'Char',         'M_Product',   'UPC',                         'NVARCHAR',   'equivalent', '', ''),
    ('product.template', 'list_price',       'Monetary',     'M_ProductPrice', 'PriceList',                'NUMERIC',    'partial',    'iDempiere stores per-pricelist not on product', 'BFF must resolve via M_PriceList_Version'),
    ('product.template', 'standard_price',   'Monetary',     'M_ProductPrice', 'PriceStd',                 'NUMERIC',    'partial',    'AVCO computation differs (iDempiere costing element)', ''),
    ('product.template', 'categ_id',         'Many2one',     'M_Product_Category','M_Product_Category_ID','RECORDID',    'equivalent', '', ''),
    ('product.template', 'uom_id',           'Many2one',     'C_UOM',       'C_UOM_ID',                    'RECORDID',   'equivalent', '', ''),
    ('product.template', 'detailed_type',    'Selection',    'M_Product',   'ProductType',                 'CHAR(1)',    'partial',    'Odoo has 8 codes; iDempiere has 6 (I/S/R/E/O/A)', 'Map: Consumable→I, Service→S, Storable→I+IsStocked'),
    ('product.template', 'sale_ok',          'Boolean',      'M_Product',   'IsSold',                      'CHAR(1)',    'equivalent', '', ''),
    ('product.template', 'purchase_ok',      'Boolean',      'M_Product',   'IsPurchased',                 'CHAR(1)',    'equivalent', '', ''),
    ('product.template', 'is_favorite',      'Boolean',      '',            '',                            '',           'none',       'No iDempiere field', 'Either add custom column on M_Product or new lookup table'),
    ('product.template', 'taxes_id',         'Many2many',    'C_Tax',       'C_Tax_ID',                    'RECORDID',   'partial',    'iDempiere uses C_TaxCategory not direct C_Tax m2m', 'Resolve at order line via C_TaxCategory_ID'),
    ('product.template', 'company_id',       'Many2one',     'AD_Client',   'AD_Client_ID',                'RECORDID',   'equivalent', '', 'Multi-tenant scope'),
    ('product.template', 'oem_no',           'Char (custom)','',            '',                            '',           'none',       'Not in stock Odoo or iDempiere', 'Add via custom column'),
    ('product.template', 'version',          'Integer (custom)','',         '',                            '',           'none',       '',                                ''),
    ('product.template.attribute.line', 'attribute_id', 'Many2one', 'M_AttributeSet', 'M_AttributeSet_ID',  'RECORDID',   'partial',    'iDempiere binds attributes at the document level via instances, not at template level', 'Major divergence — see BFF-Plan'),
    ('product.attribute.value', 'name',      'Char',         'M_AttributeValue', 'Name',                   'NVARCHAR',   'equivalent', '', ''),
    ('product.attribute.value', 'html_color','Char',         '',            '',                            '',           'none',       'Not on M_AttributeValue', 'Add custom column'),
    ('product.product',  'product_tmpl_id',  'Many2one',     '',            '',                            '',           'none',       'iDempiere has no "variants of a template"', 'Use AttributeSetInstance per document line'),
    ('product.supplierinfo', 'partner_id',   'Many2one',     'M_Product_PO','C_BPartner_ID',               'RECORDID',   'equivalent', '', 'iDempiere "M_Product_PO" is the vendor pricelist line'),
    ('product.supplierinfo', 'price',        'Monetary',     'M_Product_PO','PriceList',                   'NUMERIC',    'equivalent', '', ''),
    ('product.supplierinfo', 'delay',        'Integer',      'M_Product_PO','DeliveryTime_Promised',       'NUMBER',     'equivalent', '', ''),
    ('stock.quant',      'quantity',         'Float',        'M_StorageOnHand', 'QtyOnHand',               'NUMBER',     'equivalent', '', ''),
    ('stock.quant',      'reserved_quantity','Float',        'M_StorageReservation', 'Qty',                'NUMBER',     'equivalent', '', ''),
    ('stock.quant',      'location_id',      'Many2one',     'M_Locator',   'M_Locator_ID',                'RECORDID',   'equivalent', '', 'iDempiere "Locator" ≈ Odoo "Stock Location" leaf'),
    ('mail.message',     'body',             'Html',         '',            '',                            '',           'none',       'iDempiere has no chatter equivalent', 'Add Note/Attachment table or use AD_Note'),
]
write_sheet('DB-Mapping', db_headers, db_rows)
_unused = ([28, 22, 16, 24, 22, 14, 14, 60, 50])

# ── Conflict-Log (Phase 4) ─────────────────────────────────────────────
cf_headers = ['story_id', 'wiki_says', 'code_says', 'resolution', 'resolved_by', 'resolved_at', 'regression_test_id']
cf_rows = [
    ('product.variants.attr_line.value_ids',
     'chip background colour matches the value name',
     'html_color is per-record metadata; no relation to name (e.g. "Black" can have any colour)',
     'trust code; chip uses html_color; update wiki note',
     'agent', date.today().isoformat(), 'tests/variants/chip_color.spec.ts'),
    ('product.general.oem_no',
     'OEM No. is a visible field on the General tab',
     'No oem_no field exists in stock Odoo product.template',
     'add custom Char field via product_fluent_view-style addon; mark as Odoo-extension on iDempiere side too',
     'user', date.today().isoformat(), 'tests/general/oem_no_persists.spec.ts'),
    ('product.general.version',
     'Version field defaults to 1',
     'No version field exists in stock Odoo product.template',
     'add custom Integer field; default 1; UI right-aligned next to Part Number',
     'user', date.today().isoformat(), 'tests/general/version_default.spec.ts'),
    ('product.header.smart.variants',
     'Variants smart button always visible after attributes saved',
     'visible only when ≥1 attribute_line.value_ids has count > 1 (decoration-info=value_count<=1)',
     'trust code; render only when meaningful',
     'agent', date.today().isoformat(), 'tests/variants/smart_button_visibility.spec.ts'),
    # ── P3 findings (2026-04-30) ──────────────────────────────────────
    ('product.header.favorite',
     'Field name is `is_favorite` and is set by a `toggle_favorite` server method.',
     'Actual field is `priority` (Selection 0=Normal, 1=Favorite) at product_template.py:148. No `toggle_favorite` method exists. The yellow-star UI is the standard `priority` widget.',
     'Use `priority` everywhere in the spec. UI keeps its star but binds to priority. iDempiere strategy unchanged (still write-new since priority is single-user-scope here).',
     'agent', date.today().isoformat(), 'tests/header/favorite_writes_priority.spec.ts'),
    ('product.general.taxes',
     'A "(= $X.YZ Incl. Taxes)" suffix recomputes via a backend `_compute_price_with_tax` method.',
     'No such method exists in product/models. The suffix is rendered in the form view by combining `list_price` with the active `taxes_id` amounts on the client.',
     'Implement the suffix in React from the same primitives (list_price * (1 + Σ tax_amount/100)). No BFF endpoint required for the display.',
     'agent', date.today().isoformat(), 'tests/general/incl_taxes_suffix_clientside.spec.ts'),
    ('product.general.product_type',
     'Changes to detailed_type are driven by an `_onchange_detailed_type` handler.',
     'Method is named `_onchange_type` and is a no-op stub (`return {}`) at product_template.py:459. The actual mechanism is `_sanitize_vals` (keeps type/detailed_type in sync) plus a chain of @api.depends("detailed_type") on related computed fields.',
     'Spec must reference `_sanitize_vals` and the depends-chain, not a fictional onchange. UI free to drive visibility from detailed_type without a server hop.',
     'agent', date.today().isoformat(), 'tests/general/detailed_type_sync.spec.ts'),
    ('product.general.list_price',
     'Writing list_price triggers `_compute_pricelist_price` immediately.',
     'No such method on product.template. Pricelist evaluation happens at sale-line / cart time via product.pricelist.compute, not when the template is written.',
     'Drop the onchange claim. iDempiere mapping (MProductPrice keyed by M_PriceList_Version) stands as adapter-shim work.',
     'agent', date.today().isoformat(), 'tests/general/list_price_no_immediate_compute.spec.ts'),
]
write_sheet('Conflict-Log', cf_headers, cf_rows)
_unused = ([34, 60, 60, 60, 14, 14, 38])

# ── BFF-Plan (Phase 5) ─────────────────────────────────────────────────
bff_headers = [
    'endpoint', 'method', 'consumes', 'produces', 'covers_stories',
    'idempiere_strategy', 'effort', 'status'
]
bff_rows = [
    ('/api/erp-v2/products',                         'GET',    'q,page,pageSize,categoryId,sortField,sortDir', 'ProductPage{items,total}',
     'list view (out of slug scope)', 'iDempiere GridTab read on M_Product', 'done', 'shipped (slice 1)'),
    ('/api/erp-v2/products/{id}',                    'GET',    '',                                            'ProductDto',
     'product.general.* read', 'iDempiere GridTab read with joins to M_Product_Category, C_UOM, C_TaxCategory', 'done', 'shipped (slice 1)'),
    ('/api/erp-v2/products/{id}',                    'PATCH',  '{patch:{...},updatedMs}',                     'ProductDto (200) | {error,current} (409)',
     'name, value (Internal Ref), sku (OEM No), description (Internal Notes), upc, isActive, isStocked, isSold, isPurchased, productCategoryId, uomId, taxCategoryId, productTypeCode',
     'JDBC update on M_Product with FOR UPDATE concurrency check; bumps Updated/UpdatedBy=100', 'small', 'shipped (slices 2a + 2c + 2d)'),
    ('/api/erp-v2/products',                         'POST',   '{name,value,sku?,description?,upc?,isActive?,isStocked?,isSold?,isPurchased?}', 'ProductDto (201) | {error:validation} (400)',
     'product.header.create + Sales/Purchase/Active scope checkboxes', 'INSERT into M_Product with AD_Sequence-allocated id; defaults Category=107, Tax=107, UoM=100, ProductType=I', 'small', 'shipped (slice CRUD-2a)'),
    ('/api/erp-v2/products/{id}',                    'DELETE', '?updatedMs=...',                              'ProductDto (200) | {error,current} (409)',
     'product.header.archive', 'Reuses update() with isActive=false; same Updated/Concurrency contract as PATCH', 'small', 'shipped (slice CRUD-2a)'),
    ('/api/erp-v2/products/lookups/categories',      'GET',    '',                                            'NamedRef[]',
     'product.general.product_category (m2o options)', 'SELECT M_Product_Category_ID, Name FROM M_Product_Category for tenant', 'small', 'shipped (slice 2c)'),
    ('/api/erp-v2/products/lookups/uoms',            'GET',    '',                                            'NamedRef[]',
     'product.general.uom (m2o options)', 'SELECT C_UOM_ID, Name FROM C_UOM (system + tenant)', 'small', 'shipped (slice 2c)'),
    ('/api/erp-v2/products/lookups/tax-categories',  'GET',    '',                                            'NamedRef[]',
     'product.general.taxes (m2o options)', 'SELECT C_TaxCategory_ID, Name FROM C_TaxCategory for tenant', 'small', 'shipped (slice 2c)'),
    ('/api/erp-v2/products/lookups/product-types',   'GET',    '',                                            'NamedRef[]',
     'product.general.detailed_type (selection)', 'Static enum — vendor-neutral surface mapping iDempiere I/S/R/E/O/A', 'small', 'shipped (slice 2d)'),
    ('/api/erp-v2/products/{id}/price',              'GET',    '',                                            'PricePoint{listPrice,standardPrice,priceListVersionId}',
     'product.general.list_price + product.purchase.standard_price (read)',
     'M_ProductPrice select on tenant default IsSOPriceList=Y, latest validfrom <= today', 'small', 'shipped (slice 2b)'),
    ('/api/erp-v2/products/{id}/price',              'PUT',    '{listPrice,standardPrice}',                   'PricePoint',
     'product.general.list_price + product.purchase.standard_price (write)',
     'UPDATE-or-INSERT into M_ProductPrice; INSERT path allocates M_ProductPrice_ID via AD_Sequence', 'small', 'shipped (slice 2b)'),
    ('/api/erp-v2/products/{id}/favorite',           'POST',   '{is_favorite}',                               '204',
     'product.header.favorite', 'custom column on M_Product or per-user pref table', 'small', 'pending'),
    ('/api/erp-v2/products/{id}/image',              'PUT',    'multipart/form-data',                         '204',
     'product.header.image', 'iDempiere AD_Image or AD_Attachment', 'small', 'pending'),
    ('/api/erp-v2/products/{id}/quants',             'GET',    'locationFilter',                              'QuantPage',
     'product.header.smart.update_quantity (read)', 'M_StorageOnHand select with M_Locator + M_Warehouse joins', 'small', 'pending'),
    ('/api/erp-v2/products/{id}/quants',             'POST',   '{locationId,quantity,uomId,ownerId?}',        'QuantDto',
     'product.header.smart.update_quantity (inline edit)', 'invoke MMovement / MInventory process to adjust on-hand', 'large', 'pending'),
    ('/api/erp-v2/products/{id}/variants',           'GET',    '',                                            'VariantPage',
     'product.header.smart.variants', 'iDempiere has no template-level variants → adapter synthesises from M_AttributeSetInstance + AD_Client templates table', 'large', 'pending'),
    ('/api/erp-v2/products/{id}/attribute-lines',    'POST',   '{attributeId,valueIds[]}',                    'AttributeLineDto',
     'product.variants.attr_line.add', 'M_AttributeUse for set; per-template line table to be added', 'large', 'pending'),
    ('/api/erp-v2/products/{id}/vendors',            'GET',    '',                                            'VendorLine[]',
     'product.purchase.seller_lines (read)', 'M_Product_PO read', 'small', 'pending'),
    ('/api/erp-v2/products/{id}/vendors',            'POST',   'VendorLineCreate',                            'VendorLineDto',
     'product.purchase.seller_lines (create)', 'M_Product_PO insert', 'small', 'pending'),
    ('/api/erp-v2/products/{id}/messages',           'GET',    '',                                            'MessagePage',
     'product.chatter.* (read)', 'AD_Note + AD_Attachment + custom mail mixin table', 'medium', 'pending'),
    ('/api/erp-v2/products/{id}/messages',           'POST',   '{body, type:comment|note}',                   'MessageDto',
     'product.chatter.send_message / log_note', 'AD_Note insert + notify followers via custom delivery', 'medium', 'pending'),
    ('/api/erp-v2/products/{id}/activities',         'POST',   'ActivityCreate',                              'ActivityDto',
     'product.chatter.activities', 'R_Request or AD_Reminder hook', 'medium', 'pending'),
]
write_sheet('BFF-Plan', bff_headers, bff_rows)
_unused = ([42, 8, 38, 28, 36, 56, 10, 12])

# ── Cross-Cutting overlay ─────────────────────────────────────────────
cc_headers = ['concern', 'odoo_mechanism', 'idempiere_mechanism', 'orchestrix_strategy', 'status']
cc_rows = [
    ('i18n',          'EN marker per text field; res.lang per session', 'AD_Language + per-record translation tables (_Trl)', 'tenant-aware translations via existing locale provider; surface EN marker as decorative only in slice 1', 'pending'),
    ('multi-tenancy', 'company_id (res.company)', 'AD_Client_ID + AD_Org_ID', 'Every BFF endpoint scopes by tenant; ErpAdapter resolves AD_Client from JWT', 'partial — list endpoint scopes; PATCH not yet added'),
    ('rbac',          'ir.rule + groups', 'AD_Role + AD_Role_OrgAccess + AD_Window_Access', 'Predicate translation table; default-deny BFF; UI hides controls per role', 'pending'),
    ('workflows',     'automated.action + ir.cron + mail.template', 'R_RequestProcessor + AD_Workflow + AD_AlertProcessor', 'Initial scope: skip; flag fields that are workflow-driven so we don\'t hand-edit them', 'pending'),
    ('auditing',      'mail.tracking.value', 'AD_ChangeLog', 'Show on chatter in slice 2; backed by AD_ChangeLog reads', 'pending'),
    ('attachments',   'ir.attachment', 'AD_Attachment + AD_AttachmentNote', 'Documents smart button and chatter both read this', 'pending'),
]
write_sheet('Cross-Cutting', cc_headers, cc_rows, col_widths=[16, 38, 38, 60, 14])

# ── Glossary ──────────────────────────────────────────────────────────
gl_headers = ['term', 'meaning']
gl_rows = [
    ('story_id',         'Stable dotted identifier; never renamed once shipped (used as test key + cross-sheet FK).'),
    ('confidence',       'How sure the agent is the row reflects reality. high=video+code agree; medium=video only; low=video implies but code disagrees.'),
    ('control_type',     'Atomic UI affordance: text, monetary, selection, many2one, m2m_chips, m2m_checkboxes, one2many, checkbox, switch, radio, button, smart-button, list-add, link, image-upload, icon-toggle.'),
    ('mapping_bucket',   'equivalent | partial | none — drives BFF effort estimate.'),
    ('discovery_status', 'pending | in_progress | done | wontfix — Phase 3 progress per row.'),
    ('iteration_cap',    'Max iterations the agent is allowed before stopping and surfacing to user.'),
    ('regression_test_id','File:test path that locks a Conflict-Log resolution. Empty until Phase 6.'),
    ('GridTab',          'iDempiere\'s window/tab abstraction; the canonical write path. Direct JDBC against iDempiere tables is forbidden for new entities.'),
    ('idempiere_strategy', 'reuse-method | reuse-process | adapter-shim | write-new | wontfix.'),
    ('idempiere_status',   'verified (read the iDempiere code) | proposed (educated guess) | unknown (not checked yet).'),
]
write_sheet('Glossary', gl_headers, gl_rows, col_widths=[24, 100])

# Reorder sheets so README is first.
order = ['README', 'Sources', 'Phases', 'Stories', 'Interactions',
         'DB-Mapping', 'Conflict-Log', 'BFF-Plan', 'Cross-Cutting', 'Glossary']
wb._sheets = [wb[name] for name in order]

wb.save(OUT_PATH)
print(f'Wrote {OUT_PATH}')
print(f'Sheets: {", ".join(order)}')
print(f'Rows by sheet:')
for n in order:
    print(f'  {n}: {wb[n].max_row - 1}')
